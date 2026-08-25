"""公共基础模块（规范 7.8，先于工具 1 实现）：25 个工具唯一的基础依赖，禁止各自重新实现。

包含七个函数：read_table / check_file / ok / err / save_plot / to_jsonable / validate_columns。
本模块 import 时全局 np.random.seed(42)（规范 4 可复现性），并在 import pyplot 之前设置
matplotlib Agg 后端（附录 D 图片协议）；中文字体探测结果暴露为 CJK_FONT_OK 供绘图工具
在画图前决定中文/英文标签（红队裁决 9）。

约定：校验/读取失败一律抛 DataLabError（中文消息），工具层统一 catch 后返回 err()，
第一层工具禁止 print（红队裁决 7），日志一律走本模块 logger→stderr。
"""
from __future__ import annotations

import io
import logging
import math
import os
import re
import sys
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

# ---- 全局可复现性（规范 4）----
np.random.seed(42)

# ---- 大数据保护阈值（规范 6，红队裁决 6）----
MAX_FILE_BYTES = 50 * 1024 * 1024          # 50MB
WARN_FILE_BYTES = 5 * 1024 * 1024          # 5MB 起预检
MAX_ROWS = 2_000_000                        # 预估总行数上限
MAX_MEM_BYTES = 500 * 1024 * 1024          # 预估内存上限 500MB
SAMPLE_BYTES = 1 * 1024 * 1024             # 预检采样字节

ALLOWED_EXTS = {"csv", "tsv", "xlsx", "json"}
UNC_PREFIXES = ("\\\\", "//", "\\\\?", "\\\\.\\")
JSON_MAX_BYTES = 20 * 1024 * 1024          # json 解析放大防护（I2 红队裁决）

_PROJECT_ROOT = Path(__file__).resolve().parents[2]  # statlab_mcp/tools/_common.py -> 项目根
PLOT_DIR = _PROJECT_ROOT / "reports" / "plots"

logger = logging.getLogger("statlab_mcp")
if not logger.handlers:                     # 日志 -> stderr（红队 I3：审计轨迹落地）
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))
    logger.addHandler(_h)
    logger.setLevel(logging.INFO)

# ---- matplotlib：Agg 后端必须先行（附录 D 第 3 条），再探测中文字体 ----
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager

_CJK_CANDIDATES = ["Microsoft YaHei", "SimHei"]
CJK_FONT_OK = any(f.name in _CJK_CANDIDATES for f in font_manager.fontManager.ttflist)
if CJK_FONT_OK:
    plt.rcParams["font.sans-serif"] = _CJK_CANDIDATES
plt.rcParams["axes.unicode_minus"] = False

_WIN_RESERVED = {"con", "prn", "aux", "nul"} | {f"com{i}" for i in range(1, 10)} | {f"lpt{i}" for i in range(1, 10)}


class DataLabError(ValueError):
    """工具层业务错误：message 为面向使用者的中文提示。"""


def _norm_path(file_path: str) -> str:
    """规范化为绝对本地路径，拒绝 UNC/网络路径/空串/NUL（规范 5，红队 B/S5）。"""
    if not isinstance(file_path, str) or not file_path.strip():
        raise DataLabError("文件路径不能为空")
    if "\x00" in file_path:
        raise DataLabError("文件路径含有非法字符（NUL）")
    p = os.path.normpath(file_path.strip().strip('"'))
    if p.startswith(UNC_PREFIXES):
        raise DataLabError("仅接受本地文件路径，不支持网络路径（UNC）")
    return os.path.abspath(p)


def check_file(file_path: str) -> dict[str, Any]:
    """检查本地数据文件可用性与规模（规范 6 双重检查之一，read_table 内部自动调用）。

    拒绝：UNC 路径 / 文件不存在 / >50MB / 5-50MB 且预估行数 >200 万。
    xlsx 用 openpyxl read_only 流式取 max_row 预检（zip 压缩不能按字节估算）。
    返回 {"path": 绝对路径, "size_bytes": int, "rows_estimated": int|None}。
    """
    path = _norm_path(file_path)
    if not os.path.isfile(path):
        raise DataLabError(f"文件不存在或不可访问: {os.path.basename(path)}")
    size = os.path.getsize(path)
    if size > MAX_FILE_BYTES:
        limit_mb = MAX_FILE_BYTES // 1024 // 1024
        raise DataLabError(
            f"文件过大（{size / 1024 / 1024:.1f}MB），超过 {limit_mb}MB 上限，请拆分后重试")
    ext = os.path.splitext(path)[1].lower().lstrip(".")   # 红队 A B2：重构时丢失的定义
    rows_est = None
    if ext == "xlsx":
        # xlsx 独立分支（任意大小都预检：压缩包可声明远大于自身的内容，防 zip 炸弹，红队 I1）
        try:
            import zipfile
            with zipfile.ZipFile(path) as zf:
                total_uncompressed = sum(i.file_size for i in zf.infolist())
            if total_uncompressed > MAX_MEM_BYTES:
                raise DataLabError("Excel 解压后体积过大（疑似压缩炸弹），已拒绝")
        except DataLabError:
            raise
        except Exception:
            raise DataLabError("Excel 文件损坏或无法打开，请另存为 .xlsx 后重试") from None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            rows_est = wb.active.max_row
            wb.close()
        except Exception:
            raise DataLabError("Excel 文件损坏或无法打开，请另存为 .xlsx 后重试") from None
        if rows_est is not None and rows_est > MAX_ROWS:
            raise DataLabError(f"Excel 预估 {rows_est} 行，超过 200 万行上限，请抽样后重试")
    elif size > WARN_FILE_BYTES:
        if ext in ("csv", "tsv"):
            with open(path, "rb") as f:
                head = f.read(SAMPLE_BYTES)
            newlines = head.count(b"\n")
            if newlines == 0:
                raise DataLabError("文件前 1MB 内无换行（疑似超长单行），已拒绝解析")
            rows_est = int(newlines * size / len(head))
            if rows_est > MAX_ROWS:
                raise DataLabError(f"文件预估 {rows_est} 行，超过 200 万行上限，请抽样后重试")
        elif ext == "json":
            # json 解析放大防护（红队 I2）：展开阶段内存可能数倍于字节数
            if size > JSON_MAX_BYTES:
                raise DataLabError(f"JSON 文件超过 {JSON_MAX_BYTES // 1024 // 1024}MB，"
                                   f"解析展开可能撑爆内存，请转存为 CSV 后重试")
    return {"path": path, "size_bytes": size, "rows_estimated": rows_est}


def read_table(file_path: str) -> pd.DataFrame:
    """读表统一入口（Windows 硬性要求 2/3）：格式白名单 + 编码回退 + 内存上限兜底。

    三路分派（红队裁决 5）：
    - csv/tsv：utf-8-sig 试读 → UnicodeDecodeError 自动换 gbk → 再失败中文报错；
    - json：仅 utf-8-sig（GBK 解码几乎不失败但产出乱码，不做回退），失败报"仅支持 UTF-8"；
    - xlsx：只读第一个 sheet（openpyxl 引擎，无编码概念）。
    读取后 df.memory_usage(deep=True) 超 500MB 拒绝（红队裁决 6）。
    """
    info = check_file(file_path)
    path = info["path"]
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext not in ALLOWED_EXTS:
        raise DataLabError(f"仅支持 {sorted(ALLOWED_EXTS)} 格式，当前为 {ext or '无扩展名'}，请转换后重试")
    try:
        if ext in ("csv", "tsv"):
            sep = "\t" if ext == "tsv" else ","
            try:
                df = pd.read_csv(path, sep=sep, encoding="utf-8-sig")
            except UnicodeDecodeError:
                df = pd.read_csv(path, sep=sep, encoding="gbk")
        elif ext == "json":
            try:
                with open(path, encoding="utf-8-sig") as f:
                    raw = f.read()
            except UnicodeDecodeError:
                raise DataLabError("JSON 文件编码无法识别（仅支持 UTF-8），请另存为 UTF-8 后重试") from None
            try:
                df = pd.read_json(io.StringIO(raw))
            except ValueError:
                raise DataLabError("JSON 文件不是表格结构（需记录数组或 records 形式）") from None
        else:  # xlsx
            df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    except DataLabError:
        raise
    except pd.errors.EmptyDataError:
        raise DataLabError("文件为空或无可读数据") from None
    except UnicodeDecodeError:
        raise DataLabError("文件编码无法识别，请另存为 UTF-8 后重试") from None
    except Exception:
        logger.exception("read_table 解析失败（内部详情仅留 stderr，红队 I3）")
        raise DataLabError("文件解析失败，请检查文件内容与格式") from None
    if df.empty:
        raise DataLabError("文件为空或无可读数据")
    mem = int(df.memory_usage(deep=True).sum())
    if mem > MAX_MEM_BYTES:
        raise DataLabError(f"数据预估占用内存 {mem / 1024 / 1024:.1f}MB，超过 500MB 上限，请抽样后重试")
    return df


def to_jsonable(obj: Any) -> Any:
    """递归转换为 JSON 安全类型（规范 7.2，红队裁决 2）。

    np.float*/int*/bool_→原生；float NaN/±Inf→None；pd.Timestamp/datetime/date→str；
    np.ndarray/Series→list；dict 非 str 键→str；其余未知类型→str。
    禁止 NaN/Infinity 字面量出现在 JSON 中。
    """
    if obj is None or obj is pd.NA:
        return None
    if isinstance(obj, np.floating):  # np.floating 是 float 子类，必须最先处理
        v = float(obj)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(obj, bool):  # bool 是 int 子类，必须先判
        return bool(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, float):  # Python 原生 float 的 NaN/Inf 同样禁止进入 JSON
        v = float(obj)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(obj, int):
        return obj
    if isinstance(obj, (pd.Timestamp, datetime, date)):
        return str(obj)
    if isinstance(obj, dict):
        return {str(k): to_jsonable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [to_jsonable(v) for v in obj]
    if isinstance(obj, (np.ndarray, pd.Series)):
        return to_jsonable(obj.tolist())
    return str(obj)


def ok(data: Any, summary: str) -> dict[str, Any]:
    """统一成功包装（规范 7.1）：result 值全量过 to_jsonable，杜绝 np 类型漏网（红队裁决 8）。"""
    return {"status": "ok", "result": to_jsonable(data), "summary": str(summary)}


def err(message: str) -> dict[str, Any]:
    """统一失败包装（规范 7.1）：error 时禁止携带 result 字段。"""
    return {"status": "error", "message": str(message)}


def validate_columns(df: pd.DataFrame, required: Iterable[str]) -> None:
    """必需列存在性校验；缺失时抛中文 DataLabError。"""
    if isinstance(required, str):
        required = [required]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLabError(f"缺少必需列: {', '.join(missing)}；实际列: {list(df.columns)}")


def _estimate_period(y: pd.Series) -> int | None:
    """时序周期自动估计（FFT 主频法，设计文档 06 口径 2）。

    去线性趋势 -> rfft 幅度 -> 排除 DC 与 Nyquist -> 最大幅度对应频率的周期；
    返回 2..n/2 的整数周期，否则 None（无可估季节）。
    """
    n = int(y.size)
    x = y.to_numpy(dtype=float)
    t = np.arange(n, dtype=float)
    x = x - np.polyval(np.polyfit(t, x, 1), t)          # 去线性趋势
    spec = np.abs(np.fft.rfft(x))
    freqs = np.fft.rfftfreq(n)
    spec[0] = 0.0                                        # 去 DC
    if n % 2 == 0:
        spec[-1] = 0.0                                   # 去 Nyquist
    i = int(np.argmax(spec))
    if spec[i] <= 0 or freqs[i] <= 0:
        return None
    period = round(1.0 / freqs[i])
    return period if 2 <= period <= n // 2 else None


def _check_span_seconds(idx_min, idx_max, freq_seconds: float, what: str) -> None:
    """日期跨度防护（红队 B B1）：重采样/聚合前预估输出点数，超限拒绝（防内存爆炸）。"""
    span = (idx_max - idx_min).total_seconds()
    expected = int(span / max(float(freq_seconds), 1e-9)) + 1
    if expected > MAX_ROWS:
        raise DataLabError(
            f"日期跨度过大（{what}将生成约 {expected} 个点，超过 {MAX_ROWS} 行上限），"
            f"请拆分日期范围后重试")


def _prepare_series(df: pd.DataFrame, date_col: str, value_col: str) -> tuple:
    """时序工具公共预处理器（时序组五项统一前置，设计文档 06）。

    流程: to_datetime 容错（非法日期行剔除并计数）-> 重复时间戳按天求和聚合 ->
    频率推断（infer_freq，失败取间隔中位数）-> asfreq 固定频率 + 线性插值
    （报告插值数；两端缺失保留并注明）-> 混合时区统一 UTC。
    返回 (y: Series 索引=DatetimeIndex, meta: dict)。
    """
    if date_col not in df.columns:
        raise DataLabError(f"缺少必需列: {date_col}；实际列: {list(df.columns)}")
    if value_col not in df.columns:
        raise DataLabError(f"缺少必需列: {value_col}；实际列: {list(df.columns)}")
    if not pd.api.types.is_numeric_dtype(df[value_col]):
        raise DataLabError(f"列 {value_col} 不是数值列，无法做时序分析")

    dates = pd.to_datetime(df[date_col], errors="coerce")
    invalid = int(dates.isna().sum())
    keep = dates.notna()
    if int(keep.sum()) == 0:
        raise DataLabError("日期列无法解析")
    s = pd.Series(df.loc[keep, value_col].to_numpy(dtype=float),
                  index=dates[keep])

    # 时区统一（混合时区或 tz-aware 一律 UTC）
    utc_note = None
    if getattr(s.index, "tz", None) is not None:
        s.index = s.index.tz_convert("UTC")
        utc_note = "时区已统一为 UTC"

    # 重复时间戳 -> 按天求和聚合（规则固定为 sum）
    merged = 0
    if s.index.has_duplicates:
        _check_span_seconds(s.index.min(), s.index.max(), 86400.0, "按天聚合")
        orig = int(s.size)
        s = s.resample("D").sum()
        merged = orig - int(s.size)

    # 频率推断
    freq = pd.infer_freq(s.index)
    if freq is None:
        diffs = pd.Series(s.index).diff().dropna().dt.total_seconds()
        if diffs.size == 0:
            raise DataLabError("时间戳不足，无法推断频率")
        med_s = float(np.median(diffs))                # 以秒计数，规避 ns/us 单位混淆
        freq_offset = pd.tseries.frequencies.to_offset(pd.Timedelta(seconds=med_s))
        freq = freq_offset.freqstr                     # 仅展示用
    else:
        freq_offset = pd.tseries.frequencies.to_offset(freq)

    n_before = int(s.size)
    _check_span_seconds(s.index.min(), s.index.max(), freq_offset.nanos / 1e9, f"重采样({freq})")
    s = s.asfreq(freq_offset)
    interpolated = int(s.isna().sum())
    s = s.interpolate(method="linear")
    tail_nan = int(s.isna().sum())          # 两端缺失保留

    meta = {
        "n": int(s.size),
        "dropped_invalid_dates": invalid,
        "merged_duplicates": merged,
        "interpolated": interpolated,
        "tail_nan": tail_nan,
        "freq": str(freq),
        "utc_note": utc_note,
        "n_before_resample": n_before,
    }
    return s, meta


def _safe_name(name: str) -> str:
    """Windows 文件名清洗（红队裁决 9）：非法字符→_、去首尾空格点、保留名加前缀、限长 40。"""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip(" .") or "all"
    if name.lower() in _WIN_RESERVED:
        name = "_" + name
    return name[:40]


def save_plot(fig: Any, name: str) -> str:
    """按附录 D 保存图片，存 reports/plots/YYYYmmdd/ 子目录（红队 C B2：防堆积且不删旧图）。

    文件名 = 清洗后 name_YYYYmmdd_HHMMSS_fff.png（毫秒时间戳防同秒覆盖，红队 B S1）；
    name 由调用方拼装为 "<工具名>_<主列名或all>"；中文字体由模块顶层统一配置。
    返回图片绝对路径字符串（__image__ 字段的值）。
    """
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S_%f")[:-3]
    day_dir = PLOT_DIR / now.strftime("%Y%m%d")
    day_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{_safe_name(name)}_{ts}.png"
    out = day_dir / fname
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return str(out)
