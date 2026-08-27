"""公共基础模块（规范 7.8，先于工具 1 实现）：25 个工具唯一的基础依赖，禁止各自重新实现。

包含七个函数：read_table / check_file / ok / err / save_plot / to_jsonable / validate_columns。
本模块 import 时全局 np.random.seed(42)（规范 4 可复现性），并在 import pyplot 之前设置
matplotlib Agg 后端（附录 D 图片协议）；中文字体探测结果暴露为 CJK_FONT_OK 供绘图工具
在画图前决定中文/英文标签（红队裁决 9）。

约定：校验/读取失败一律抛 DataLabError（中文消息），工具层统一 catch 后返回 err()，
第一层工具禁止 print（红队裁决 7），日志一律走本模块 logger→stderr。
"""
from __future__ import annotations

import hashlib
import io
import logging
import math
import os
import re
import shutil
import sys
import threading
from collections import OrderedDict
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

# ---- 全局可复现性（规范 4）----
np.random.seed(42)

# ---- 大数据保护阈值（规范 6，红队裁决 6）----
MAX_FILE_BYTES = 50 * 1024 * 1024          # 50MB
WARN_FILE_BYTES = 5 * 1024 * 1024          # 5MB 起预检
MAX_ROWS = 2_000_000                        # 预估总行数上限
MAX_MEM_BYTES = 500 * 1024 * 1024          # 预估内存上限 500MB
SAMPLE_BYTES = 1 * 1024 * 1024             # 预检采样字节

ALLOWED_EXTS = {"csv", "tsv", "xlsx", "json"}
UNC_PREFIXES = ("\\\\", "//", "\\\\?", "\\\\.\\")
JSON_MAX_BYTES = 20 * 1024 * 1024          # json 解析放大防护（I2 红队裁决）

_PROJECT_ROOT = Path(__file__).resolve().parents[2]  # statlab_mcp/tools/_common.py -> 项目根
PLOT_DIR = _PROJECT_ROOT / "reports" / "plots"
IMPUTED_DIR = _PROJECT_ROOT / "reports" / "imputed"   # v1.2.0 T1 插补产物目录（SPEC 第 11 节）

logger = logging.getLogger("statlab_mcp")
if not logger.handlers:                     # 日志 -> stderr（红队 I3：审计轨迹落地）
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))
    logger.addHandler(_h)
    logger.setLevel(logging.INFO)

# ---- matplotlib：Agg 后端必须先行（附录 D 第 3 条），再探测中文字体 ----
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager

_CJK_CANDIDATES = ["Microsoft YaHei", "SimHei"]
CJK_FONT_OK = any(f.name in _CJK_CANDIDATES for f in font_manager.fontManager.ttflist)
if CJK_FONT_OK:
    plt.rcParams["font.sans-serif"] = _CJK_CANDIDATES
plt.rcParams["axes.unicode_minus"] = False

_WIN_RESERVED = {"con", "prn", "aux", "nul"} | {f"com{i}" for i in range(1, 10)} | {f"lpt{i}" for i in range(1, 10)}


class DataLabError(ValueError):
    """工具层业务错误：message 为面向使用者的中文提示，code 为机器可读错误码（SPEC 第 9 节）。

    code 一经发布永久稳定、只增不改不复用；DataLabError 单参构造时默认兜底码 E9999。
    """

    def __init__(self, message: str, code: str = "E9999") -> None:
        super().__init__(message)
        self.code = code


class EC:
    """错误码表（与 statlab_mcp/docs/SPEC.md 第 9 节一致，check_readme_claims 扩展项核对两边一致）。

    码格式：E + 四位数字。语义详见 SPEC；此处仅集中定义，杜绝工具层散落字符串。
    """

    PARAM = "E1001"           # 参数校验失败（含 pydantic 层）
    PATH = "E1002"            # 路径非法（空/UNC/含 NUL 等）
    FILE_MISSING = "E1003"    # 文件不存在或不可访问
    FILE_EMPTY = "E1004"      # 文件为空或无可读数据
    SCALE = "E1005"           # 文件/数据规模超限（体积/行数/内存/日期跨度/JSON 放大/zip 炸弹）
    FORMAT = "E1006"          # 文件格式不支持
    ENCODING = "E1007"        # 文件编码无法识别
    COLUMN_MISSING = "E1008"  # 缺少必需列
    COLUMN_TYPE = "E1009"     # 列非数值（或列类型不符合工具要求）
    INSUFFICIENT = "E1010"    # 样本量/有效值不足
    STRUCTURE = "E1011"       # 分组/配对结构非法（组数不符、配对无变异等）
    NO_OBJECT = "E1012"       # 数据无可处理对象（无须处理，如无缺失可插补；v1.2.0 起）
    CALC = "E9999"            # 计算失败兜底


def _norm_path(file_path: str) -> str:
    """规范化为绝对本地路径，拒绝 UNC/网络路径/空串/NUL（规范 5，红队 B/S5）。"""
    if not isinstance(file_path, str) or not file_path.strip():
        raise DataLabError("文件路径不能为空", EC.PATH)
    if "\x00" in file_path:
        raise DataLabError("文件路径含有非法字符（NUL）", EC.PATH)
    p = os.path.normpath(file_path.strip().strip('"'))
    if p.startswith(UNC_PREFIXES):
        raise DataLabError("仅接受本地文件路径，不支持网络路径（UNC）", EC.PATH)
    return os.path.abspath(p)


def check_file(file_path: str) -> dict[str, Any]:
    """检查本地数据文件可用性与规模（规范 6 双重检查之一，read_table 内部自动调用）。

    拒绝：UNC 路径 / 文件不存在 / >50MB / 5-50MB 且预估行数 >200 万。
    xlsx 用 openpyxl read_only 流式取 max_row 预检（zip 压缩不能按字节估算）。
    返回 {"path": 绝对路径, "size_bytes": int, "rows_estimated": int|None}。
    """
    path = _norm_path(file_path)
    if not os.path.isfile(path):
        raise DataLabError(f"文件不存在或不可访问: {os.path.basename(path)}", EC.FILE_MISSING)
    size = os.path.getsize(path)
    if size > MAX_FILE_BYTES:
        limit_mb = MAX_FILE_BYTES // 1024 // 1024
        raise DataLabError(
            f"文件过大（{size / 1024 / 1024:.1f}MB），超过 {limit_mb}MB 上限，请拆分后重试", EC.SCALE)
    ext = os.path.splitext(path)[1].lower().lstrip(".")   # 红队 A B2：重构时丢失的定义
    rows_est = None
    if ext == "xlsx":
        # xlsx 独立分支（任意大小都预检：压缩包可声明远大于自身的内容，防 zip 炸弹，红队 I1）
        try:
            import zipfile
            with zipfile.ZipFile(path) as zf:
                total_uncompressed = sum(i.file_size for i in zf.infolist())
            if total_uncompressed > MAX_MEM_BYTES:
                raise DataLabError("Excel 解压后体积过大（疑似压缩炸弹），已拒绝", EC.SCALE)
        except DataLabError:
            raise
        except Exception:
            raise DataLabError("Excel 文件损坏或无法打开，请另存为 .xlsx 后重试", EC.FILE_EMPTY) from None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            rows_est = wb.active.max_row
            wb.close()
        except Exception:
            raise DataLabError("Excel 文件损坏或无法打开，请另存为 .xlsx 后重试", EC.FILE_EMPTY) from None
        if rows_est is not None and rows_est > MAX_ROWS:
            raise DataLabError(f"Excel 预估 {rows_est} 行，超过 200 万行上限，请抽样后重试", EC.SCALE)
    elif size > WARN_FILE_BYTES:
        if ext in ("csv", "tsv"):
            with open(path, "rb") as f:
                head = f.read(SAMPLE_BYTES)
            newlines = head.count(b"\n")
            if newlines == 0:
                raise DataLabError("文件前 1MB 内无换行（疑似超长单行），已拒绝解析", EC.SCALE)
            rows_est = int(newlines * size / len(head))
            if rows_est > MAX_ROWS:
                raise DataLabError(f"文件预估 {rows_est} 行，超过 200 万行上限，请抽样后重试", EC.SCALE)
        elif ext == "json":
            # json 解析放大防护（红队 I2）：展开阶段内存可能数倍于字节数
            if size > JSON_MAX_BYTES:
                raise DataLabError(f"JSON 文件超过 {JSON_MAX_BYTES // 1024 // 1024}MB，"
                                   f"解析展开可能撑爆内存，请转存为 CSV 后重试", EC.SCALE)
    return {"path": path, "size_bytes": size, "rows_estimated": rows_est}


def read_table(file_path: str) -> pd.DataFrame:
    """读表统一入口（Windows 硬性要求 2/3）：格式白名单 + 编码回退 + 内存上限兜底。

    三路分派（红队裁决 5）：
    - csv/tsv：utf-8-sig 试读 → UnicodeDecodeError 自动换 gbk → 再失败中文报错；
    - json：仅 utf-8-sig（GBK 解码几乎不失败但产出乱码，不做回退），失败报"仅支持 UTF-8"；
    - xlsx：只读第一个 sheet（openpyxl 引擎，无编码概念）。
    读取后 df.memory_usage(deep=True) 超 500MB 拒绝（红队裁决 6）。
    v1.1.0 P1-1：解析结果经两级键进程内 LRU 缓存（廉价键=路径归一化+大小+mtime，
    命中后以 SHA256 验内容），见 _file_cache_get/_file_cache_put；防护全部先于缓存执行；
    缓存命中返回共享引用依赖 pandas>=3 CoW 语义（锁定版本即 ≥3，规范 8）。
    """
    info = check_file(file_path)                       # 防护 1/2：路径与体积/行数预检
    path = info["path"]
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext not in ALLOWED_EXTS:
        raise DataLabError(f"仅支持 {sorted(ALLOWED_EXTS)} 格式，当前为 {ext or '无扩展名'}，请转换后重试", EC.FORMAT)

    cached = _file_cache_get(path)
    if cached is not None:
        return cached
    df = _parse_table(path, ext)
    _file_cache_put(path, df)
    return df


# ---- 进程内文件缓存（v1.1.0 P1-1；规则钉死，SPEC 第 4 节）----

NO_CACHE_ENV = "STATLAB_NO_CACHE"
_CACHE_MAX_ENTRIES = 8                               # 容量上限：条目数
_CACHE_MAX_MEM = MAX_MEM_BYTES                       # 总内存预算 500MB（同防护口径）

_cache_lock = threading.Lock()
_cache_cheap: OrderedDict[tuple, dict] = OrderedDict()   # 廉价键 -> 条目（LRU 序）


def _cache_disabled_by_env(stream: Any = None) -> bool:
    """STATLAB_NO_CACHE=1 时禁用缓存（测试对照专用，生产文档不宣传）；非法取值告警并忽略。"""
    v = os.environ.get(NO_CACHE_ENV)
    if v is None or v == "":
        return False
    if v == "1":
        return True
    print(f"[statlab-mcp] 告警：环境变量 {NO_CACHE_ENV}={v!r} 非法（仅支持 '1' 表示禁用），"
          "已忽略该设置", file=stream or sys.stderr)
    return False


def _sha256_of_file(path: str) -> str:
    """流式计算文件 SHA256（不整载内存）。"""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _cheap_key(path: str) -> tuple:
    """一级廉价键：（normcase 归一化绝对路径, 大小, mtime ns）——查询不读全文件。"""
    p = Path(path)
    st = p.stat()
    return (os.path.normcase(str(p.resolve())), st.st_size, st.st_mtime_ns)


def _file_cache_get(path: str) -> pd.DataFrame | None:
    """两级键查询：廉价键命中后再验 SHA256，防 mtime 被伪造/精度丢失；未命中不读全文件。

    锁内只碰字典，SHA256 在锁外流式计算；比对期间条目被替换/淘汰则按未命中处理。
    """
    if _cache_disabled_by_env():
        return None
    try:
        key = _cheap_key(path)
    except OSError:
        return None                                   # 文件不可 stat：交给后续正常报错
    with _cache_lock:
        entry = _cache_cheap.get(key)
        if entry is None:
            return None
        wanted = entry["sha256"]
    current = _sha256_of_file(path)
    with _cache_lock:
        if _cache_cheap.get(key) is not entry:
            return None                               # 已被覆盖/淘汰：视为未命中
        if current != wanted:
            _cache_cheap.pop(key, None)               # 内容已变：淘汰并按未命中处理
            return None
        _cache_cheap.move_to_end(key)                 # LRU 提位
        return entry["df"]


def _file_cache_put(path: str, df: pd.DataFrame) -> None:
    """插入前按预算先淘汰（LRU 顺序）；单条自身超预算则不缓存；不持久化、不加盘。"""
    if _cache_disabled_by_env():
        return
    est = int(df.memory_usage(deep=True).sum())
    if est > _CACHE_MAX_MEM:
        logger.info("文件 %s 数据量 %.0fMB 超缓存预算 %dMB，本次结果不进缓存",
                    os.path.basename(path), est / 1024 / 1024,
                    _CACHE_MAX_MEM // 1024 // 1024)
        return
    key = _cheap_key(path)
    try:
        digest = _sha256_of_file(path)                # 锁外计算，缩短临界区
    except OSError:
        # 竞态窗口（红队 P2-2）：解析成功后文件消失——缓存写入尽力而为，
        # 静默放弃与 _file_cache_get 的 OSError 防护对称，不影响本次返回
        return
    with _cache_lock:
        _cache_cheap.pop(key, None)                   # 同键覆盖为最新引用
        while len(_cache_cheap) >= _CACHE_MAX_ENTRIES or \
                sum(e["bytes"] for e in _cache_cheap.values()) + est > _CACHE_MAX_MEM:
            if not _cache_cheap:
                break
            _cache_cheap.popitem(last=False)          # 最老优先淘汰
        _cache_cheap[key] = {"df": df, "bytes": est, "sha256": digest}


def _parse_table(path: str, ext: str) -> pd.DataFrame:
    """按格式三路分派解析（仅供 read_table；路径/规模防护与白名单均在上游执行）。

    - csv/tsv：utf-8-sig 试读 → UnicodeDecodeError 自动换 gbk → 再失败中文报错；
    - json：仅 utf-8-sig（GBK 静默乱码比报错更危险，不做回退）；
    - xlsx：openpyxl 引擎只读第一个 sheet（无编码概念）。
    """
    try:
        if ext in ("csv", "tsv"):
            sep = "\t" if ext == "tsv" else ","
            try:
                df = pd.read_csv(path, sep=sep, encoding="utf-8-sig")
            except UnicodeDecodeError:
                df = pd.read_csv(path, sep=sep, encoding="gbk")
            # 重复列名检测（pandas 已自动改名 x -> x.1，须如实注明；外部评审 L10）
            try:
                # 与 read_csv 相同的编码链（utf-8-sig 失败换 gbk）：固定 errors="replace"
                # 会把 GBK 字节替换成 U+FFFD，不同中文列名可能误判为重复（红队复检新发现 1）
                enc = "utf-8-sig"
                try:
                    with open(path, encoding=enc) as f:
                        header_line = f.readline()
                except UnicodeDecodeError:
                    enc = "gbk"
                    with open(path, encoding=enc) as f:
                        header_line = f.readline()
                raw_cols = header_line.rstrip("\r\n").split(sep)
                # 仅当原始表头确有重复且 pandas 已实际改名（df.columns != raw_cols）才记录：
                # 无表头文件的首行是数据行，不能据此误报"重复列名"
                if any(raw_cols.count(c) > 1 for c in raw_cols) \
                        and [str(c) for c in df.columns] != raw_cols:
                    dups = sorted({c for c in raw_cols if raw_cols.count(c) > 1})
                    df.attrs["duplicate_columns_renamed"] = dups
            except Exception:
                pass   # 检测失败不影响读表
        elif ext == "json":
            try:
                with open(path, encoding="utf-8-sig") as f:
                    raw = f.read()
            except UnicodeDecodeError:
                raise DataLabError(
                    "JSON 文件编码无法识别（仅支持 UTF-8），请另存为 UTF-8 后重试",
                    EC.ENCODING) from None
            try:
                df = pd.read_json(io.StringIO(raw))
            except ValueError:
                raise DataLabError("JSON 文件不是表格结构（需记录数组或 records 形式）", EC.FORMAT) from None
        else:  # xlsx
            df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    except DataLabError:
        raise
    except pd.errors.EmptyDataError:
        raise DataLabError("文件为空或无可读数据", EC.FILE_EMPTY) from None
    except UnicodeDecodeError:
        raise DataLabError("文件编码无法识别，请另存为 UTF-8 后重试", EC.ENCODING) from None
    except Exception:
        logger.exception("read_table 解析失败（内部详情仅留 stderr，红队 I3）")
        raise DataLabError("文件解析失败，请检查文件内容与格式", EC.CALC) from None
    if df.empty:
        raise DataLabError("文件为空或无可读数据", EC.FILE_EMPTY)
    mem = int(df.memory_usage(deep=True).sum())
    if mem > MAX_MEM_BYTES:
        raise DataLabError(f"数据预估占用内存 {mem / 1024 / 1024:.1f}MB，超过 500MB 上限，请抽样后重试", EC.SCALE)
    return df


# ---- inline 小数据通道（v1.2.0 T3；协议契约见 SPEC 第 12 节）----

_INLINE_MAX_ROWS = 10_000
_INLINE_MAX_COLS = 200
_INLINE_MAX_CELLS = 50_000
_INLINE_MAX_PAYLOAD = 16 * 1024 * 1024          # 序列化总字节（文件通道 50MB 的对称防线）
_INLINE_MAX_CELL_CHARS = 65_536
_CTRL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")   # 控制字符清洗


def resolve_data(file_path: str | None, inline_data: Any,
                 *, require_input: bool = True) -> tuple[pd.DataFrame, str]:
    """file_path / inline_data 二选一单点分派（T3；禁止各工具自行解析）。

    返回 (DataFrame, data_source)，data_source ∈ {"file", "inline"}；
    analysis_plan 等可选源工具传 require_input=False：双缺返回 (空 DataFrame, "none")。
    """
    # 红队 P1-1（C13b）：空串/空白 file_path 不算"无来源"——它必须沿文件通道
    # 走 read_table→_norm_path 报 E1002"文件路径不能为空"（v1.0.3 旧场景文案
    # 与错误码红线），而非被吞成 E1001"缺少数据来源"
    has_file = file_path is not None
    has_inline = inline_data is not None
    if has_file and has_inline:
        raise DataLabError("file_path 与 inline_data 只能提供其一，请选择一种数据来源",
                           EC.PARAM)
    if not has_file and not has_inline:
        if require_input:
            raise DataLabError(
                "缺少数据来源：请提供 file_path（数据文件路径）或 "
                "inline_data（内联小数据）之一", EC.PARAM)
        return pd.DataFrame(), "none"
    if has_file:
        return read_table(str(file_path)), "file"
    return normalize_inline(inline_data), "inline"


def _normalize_dtype(df: pd.DataFrame) -> pd.DataFrame:
    """inline 构造后的 dtype 归一（SPEC §12.4）：全 null 列→float64、数值混合→float64。"""
    for col in df.columns:
        non_null = df[col].dropna()
        if non_null.empty:
            df[col] = df[col].astype("float64")
            continue
        if all(isinstance(v, (int, float)) and not isinstance(v, bool)
               for v in non_null):
            df[col] = df[col].astype("float64")
        # 纯 bool / 含 str 的列保持 object，由各工具按自身口径处理
    return df


def _validate_cell(v: Any) -> Any:
    """单元格校验：值域标量 + 单 cell 长度上限（E1001/E1005），返回规范化值。

    bool 判序先于 int；NaN/Inf 作为数据值放行（与文件读取口径一致，
    输出侧统一 to_jsonable 转 null——铁律 6 约束输出而非输入）。
    """
    if v is None or isinstance(v, (bool, int, float, str)):
        if isinstance(v, str):
            if len(v) > _INLINE_MAX_CELL_CHARS:
                raise DataLabError(
                    f"inline 单元格字符串长度 {len(v)} 超过 {_INLINE_MAX_CELL_CHARS} 上限；"
                    "请落盘后改用 file_path", EC.SCALE)
            cleaned = _CTRL_CHARS_RE.sub("", v)
            return cleaned if cleaned != v else v
        return v
    raise DataLabError(
        f"inline 单元格不允许 {type(v).__name__} 类型（嵌套结构/对象均不支持）；"
        "仅支持字符串/数值/布尔/null 标量", EC.PARAM)


def normalize_inline(inline_data: Any) -> pd.DataFrame:
    """inline 数据归一化入口：形态识别 → 五项规模上限 → 值域校验 → dtype 归一。

    实现纪律：单遍扫描在构造 DataFrame **之前**完成全部拒绝判定（防大对象构造后拒）。
    """
    # ---- 形态识别：split dict（header+rows）或 records list（键并集为列序）----
    header: list[str] | None = None
    rows_raw: list[list[Any]] = []
    records: list[dict[str, Any]] | None = None

    if isinstance(inline_data, dict):
        if "header" not in inline_data or "rows" not in inline_data:
            raise DataLabError(
                "inline dict 形态必须同时包含 header 与 rows 字段", EC.PARAM)
        header_raw = inline_data["header"]
        raw_rows = inline_data["rows"]
        if not isinstance(header_raw, list) or \
                not all(isinstance(h, str) for h in header_raw):
            raise DataLabError("inline split 形态的 header 必须为字符串数组", EC.PARAM)
        if len(set(header_raw)) != len(header_raw):
            dup = sorted({h for h in header_raw if header_raw.count(h) > 1})
            raise DataLabError(
                f"inline split 形态 header 存在重复列名：{dup}；"
                "请改用唯一列名（与文件通道重复列名处理口径对齐）", EC.PARAM)
        if not isinstance(raw_rows, list):
            raise DataLabError("inline split 形态的 rows 必须为数组", EC.PARAM)
        for row in raw_rows:
            if not isinstance(row, list) or len(row) != len(header_raw):
                raise DataLabError(
                    "split 形态每行长度必须等于 header 长度"
                    f"（期望 {len(header_raw)} 个值）", EC.PARAM)
        # 行数上限先行（P2-5：先拒绝后拷贝，防御纵深）
        if len(raw_rows) > _INLINE_MAX_ROWS:
            raise DataLabError(
                f"inline 数据 {len(raw_rows)} 行超过 {_INLINE_MAX_ROWS} 行上限；"
                "请落盘后改用 file_path", EC.SCALE)
        header = list(header_raw)
        rows_raw = [list(r) for r in raw_rows]
    elif isinstance(inline_data, list):
        keys: list[str] = []
        seen: set[str] = set()
        for rec in inline_data:
            if not isinstance(rec, dict):
                raise DataLabError(
                    "records 形态元素必须为 {'列名': 值} 字典（数组第 "
                    f"{len(keys)} 行之后发现非对象）", EC.PARAM)
            for k in rec:
                if not isinstance(k, str):
                    raise DataLabError("records 形态的列名必须为字符串", EC.PARAM)
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
        header = keys                        # 列集取各行键的并集（保持首现顺序）
        records = inline_data
        # 行值延后按最终并集逐列取值（rec.get 自动补 null），无需中间对齐
    else:
        raise DataLabError(
            "inline_data 仅支持 records 数组或 "
            "{'header': [...], 'rows': [[...], ...]} 对象形态", EC.PARAM)

    # ---- 规模上限：行/列/单元格三查（payload 总量与单格长度在逐格校验中）----
    n_cols = len(header) if header is not None else 0
    n_rows = len(rows_raw) if records is None else len(records)
    if n_rows == 0 or n_cols == 0:
        raise DataLabError("inline 数据为空（无行或无列），无可分析内容", EC.FILE_EMPTY)
    if n_cols > _INLINE_MAX_COLS:
        raise DataLabError(f"inline 数据 {n_cols} 列超过 {_INLINE_MAX_COLS} 列上限；"
                           "请精简列后重试", EC.SCALE)
    if n_rows > _INLINE_MAX_ROWS:
        raise DataLabError(
            f"inline 数据 {n_rows} 行超过 {_INLINE_MAX_ROWS} 行上限；"
            "请落盘后改用 file_path", EC.SCALE)
    if n_rows * n_cols > _INLINE_MAX_CELLS:
        raise DataLabError(
            f"inline 数据共 {n_rows * n_cols} 个单元格超过 {_INLINE_MAX_CELLS} 上限；"
            "请落盘为文件后改用 file_path", EC.SCALE)

    # ---- 逐格值域校验 + payload 累计（单遍；不通过即抛，绝不清洗放行）----
    payload_total = 0
    cells_out: list[list[Any]] = []
    if records is None:                      # split 形态
        for row in rows_raw:
            validated_row = []
            for v in row:
                v2 = _validate_cell(v)
                if isinstance(v2, str):
                    payload_total += len(v2.encode("utf-8", errors="replace"))
                elif isinstance(v2, (int, float)) and not isinstance(v2, bool):
                    payload_total += 8
                validated_row.append(v2)
            cells_out.append(validated_row)
    else:                                     # records 形态：按最终并集逐格取值
        for rec in records:
            validated_row = []
            for k in header or []:
                v2 = _validate_cell(rec.get(k))
                if isinstance(v2, str):
                    payload_total += len(v2.encode("utf-8", errors="replace"))
                elif isinstance(v2, (int, float)) and not isinstance(v2, bool):
                    payload_total += 8
                validated_row.append(v2)
            cells_out.append(validated_row)
    if payload_total > _INLINE_MAX_PAYLOAD:
        raise DataLabError(
            f"inline 数据序列化总量约 {payload_total // 1024 // 1024}MB 超过 16MB 上限；"
            "请落盘为文件后改用 file_path", EC.SCALE)

    # ---- 构造 DataFrame + dtype 归一（int64/float64/object 三态收口）----
    df = pd.DataFrame(cells_out, columns=header or [], dtype=object)
    df = _normalize_dtype(df)
    if df.empty:
        raise DataLabError("inline 数据为空（无行或无列），无可分析内容", EC.FILE_EMPTY)
    return df


def require_non_none(**named: Any) -> None:
    """运行期必填参数强校验（T3 连锁 optional 化的契约补偿，SPEC §12.6）。

    背景：file_path 转可选后，Python 语法强制其后的既有必填参数同样带默认值，
    JSON Schema 的 required 全集随之消失；本函数在函数体入口以精确中文逐名校验，
    保证"漏传即报错"的 agent 体验与 schema required 时代完全等价。
    """
    missing = [k for k, v in named.items() if v is None]
    if missing:
        raise DataLabError(f"缺少必需参数: {', '.join(missing)}", EC.PARAM)


def to_jsonable(obj: Any) -> Any:
    """递归转换为 JSON 安全类型（规范 7.2，红队裁决 2）。

    np.float*/int*/bool_→原生；float NaN/±Inf→None；pd.Timestamp/datetime/date→str；
    np.ndarray/Series→list；dict 非 str 键→str；其余未知类型→str。
    禁止 NaN/Infinity 字面量出现在 JSON 中。
    """
    if obj is None or obj is pd.NA or obj is pd.NaT:
        return None                                  # 红队 P2-1：NaT 与 NA 同归 null
    if isinstance(obj, np.floating):  # np.floating 是 float 子类，必须最先处理
        v = float(obj)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(obj, bool):  # bool 是 int 子类，必须先判
        return bool(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, float):  # Python 原生 float 的 NaN/Inf 同样禁止进入 JSON
        v = float(obj)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(obj, int):
        return obj
    if isinstance(obj, (pd.Timestamp, datetime, date)):
        return str(obj)
    if isinstance(obj, dict):
        return {str(k): to_jsonable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [to_jsonable(v) for v in obj]
    if isinstance(obj, (np.ndarray, pd.Series)):
        return to_jsonable(obj.tolist())
    return str(obj)


def ok(data: Any, summary: str) -> dict[str, Any]:
    """统一成功包装（规范 7.1）：result 值全量过 to_jsonable，杜绝 np 类型漏网（红队裁决 8）。"""
    return {"status": "ok", "result": to_jsonable(data), "summary": str(summary)}


def err(code: str, message: str) -> dict[str, Any]:
    """统一失败包装（规范 7.1，v1.1.0 起错误码必填）：error 时禁止携带 result 字段。

    code 必填是为杜绝漏码；调用方优先透传 DataLabError.code（工具层模板 err(e.code, ...)），
    计算兜底场景用 EC.CALC。
    """
    return {"status": "error", "error_code": str(code), "message": str(message)}


def validate_columns(df: pd.DataFrame, required: Iterable[str]) -> None:
    """必需列存在性校验；缺失时抛中文 DataLabError。"""
    if isinstance(required, str):
        required = [required]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise DataLabError(
            f"缺少必需列: {', '.join(missing)}；实际列: {list(df.columns)}",
            EC.COLUMN_MISSING)


def _estimate_period(y: pd.Series) -> int | None:
    """时序周期自动估计（FFT 主频法，设计文档 06 口径 2）。

    去线性趋势 -> rfft 幅度 -> 排除 DC 与 Nyquist -> 最大幅度对应频率的周期；
    返回 2..n/2 的整数周期，否则 None（无可估季节）。
    """
    n = int(y.size)
    x = y.to_numpy(dtype=float)
    t = np.arange(n, dtype=float)
    x = x - np.polyval(np.polyfit(t, x, 1), t)          # 去线性趋势
    spec = np.abs(np.fft.rfft(x))
    freqs = np.fft.rfftfreq(n)
    spec[0] = 0.0                                        # 去 DC
    if n % 2 == 0:
        spec[-1] = 0.0                                   # 去 Nyquist
    i = int(np.argmax(spec))
    if spec[i] <= 0 or freqs[i] <= 0:
        return None
    period = round(1.0 / freqs[i])
    return period if 2 <= period <= n // 2 else None


def _check_span_seconds(idx_min, idx_max, freq_seconds: float, what: str) -> None:
    """日期跨度防护（红队 B B1）：重采样/聚合前预估输出点数，超限拒绝（防内存爆炸）。"""
    span = (idx_max - idx_min).total_seconds()
    expected = int(span / max(float(freq_seconds), 1e-9)) + 1
    if expected > MAX_ROWS:
        raise DataLabError(
            f"日期跨度过大（{what}将生成约 {expected} 个点，超过 {MAX_ROWS} 行上限），"
            f"请拆分日期范围后重试", EC.SCALE)


def _prepare_series(df: pd.DataFrame, date_col: str, value_col: str) -> tuple:
    """时序工具公共预处理器（时序组五项统一前置，设计文档 06）。

    流程: to_datetime 容错（非法日期行剔除并计数）-> 重复时间戳按天求和聚合 ->
    频率推断（infer_freq，失败取间隔中位数）-> asfreq 固定频率 + 线性插值
    （报告插值数；两端缺失保留并注明）-> 混合时区统一 UTC。
    返回 (y: Series 索引=DatetimeIndex, meta: dict)。
    """
    if date_col not in df.columns:
        raise DataLabError(f"缺少必需列: {date_col}；实际列: {list(df.columns)}", EC.COLUMN_MISSING)
    if value_col not in df.columns:
        raise DataLabError(f"缺少必需列: {value_col}；实际列: {list(df.columns)}", EC.COLUMN_MISSING)
    if not pd.api.types.is_numeric_dtype(df[value_col]):
        raise DataLabError(f"列 {value_col} 不是数值列，无法做时序分析", EC.COLUMN_TYPE)

    # utc=True：混合时区输入在此统一 UTC（红队 P1-2：coerce 不覆盖 Mixed timezones
    # ValueError，原实现会冒泡 E9999）；单一 tz-aware 序列语义不变
    dates = pd.to_datetime(df[date_col], errors="coerce", utc=True)
    invalid = int(dates.isna().sum())
    keep = dates.notna()
    if int(keep.sum()) == 0:
        raise DataLabError("日期列无法解析", EC.INSUFFICIENT)
    # sort_index：乱序/倒序输入先排序（红队 P1-1：乱序会让 infer_freq/跨度检查
    # 产生病态负步长，误报 E1005"日期跨度过大"）
    s = pd.Series(df.loc[keep, value_col].to_numpy(dtype=float),
                  index=dates[keep]).sort_index()

    # 时区统一：utc=True 使 naive / tz-aware / 混合时区一律落 UTC（红队 P1-2：
    # 混合时区此前 coerce 不覆盖 Mixed timezones ValueError，会冒泡 E9999，
    # "混合时区统一 UTC"承诺未兑现——现兑现）。注记仅在原始数据自带时区时出现
    # （naive 输入不产生 summary 尾巴，输出逐字节不变；抽样 3 个，漏标方向保守）
    utc_note = None
    for _v in df.loc[keep, date_col].head(3):
        if isinstance(_v, str) and re.search(r"(?:[+-]\d{2}:?\d{2}|Z)\s*$", _v):
            utc_note = "时区已统一为 UTC"
            break
        if getattr(_v, "tzinfo", None) is not None:
            utc_note = "时区已统一为 UTC"
            break

    # 重复时间戳 -> 按天求和聚合（规则固定为 sum）
    merged = 0
    dup_note = None
    if s.index.has_duplicates:
        _check_span_seconds(s.index.min(), s.index.max(), 86400.0, "按天聚合")
        orig = int(s.size)
        # 日内频率检测（外部评审 M1）：若原始含时分秒，按天聚合会损失日内粒度，须如实说明
        intraday = bool(np.any(np.asarray(s.index.hour) > 0)
                        or np.any(np.asarray(s.index.minute) > 0)
                        or np.any(np.asarray(s.index.second) > 0))
        # min_count=1：全缺失日保持 NaN 而非 0（红队 P1-5：sum 的 skipna 语义会把
        # 全 NaN 组写成 0.0，静默伪造观测值；NaN 走既有插值/披露路径）
        s = s.resample("D").sum(min_count=1)
        merged = orig - int(s.size)
        if intraday:
            dup_note = ("原始数据为日内频率（含时分秒），因存在重复时间戳，"
                        "已整体按天求和聚合，日内粒度已损失")

    # 频率推断
    freq = pd.infer_freq(s.index)
    if freq is None:
        diffs = pd.Series(s.index).diff().dropna().dt.total_seconds()
        if diffs.size == 0:
            raise DataLabError("时间戳不足，无法推断频率", EC.INSUFFICIENT)
        med_s = float(np.median(diffs))                # 以秒计数，规避 ns/us 单位混淆
        freq_offset = pd.tseries.frequencies.to_offset(pd.Timedelta(seconds=med_s))
        freq = freq_offset.freqstr                     # 仅展示用
    else:
        freq_offset = pd.tseries.frequencies.to_offset(freq)

    n_before = int(s.size)
    # 频率步长（秒）：以锚点实测一步的真实时长，兼容 MonthEnd/QuarterEnd/YearEnd 等
    # 非固定频率（其 .nanos 在 pandas 3.x 直接抛 ValueError，且旧版返回 0 会误判超限；
    # 外部评审 S1 修复：月/季/年频此前 100% 失效）
    step_td = (s.index.min() + freq_offset) - s.index.min()
    _check_span_seconds(s.index.min(), s.index.max(),
                        max(float(step_td.total_seconds()), 1e-9), f"重采样({freq})")
    s = s.asfreq(freq_offset)
    gap_total = int(s.isna().sum())
    # limit_area="inside"：仅插值两端有效值之间的缺失；头部/尾部缺失一律保留
    # （红队复检：pandas 默认 forward 方向会用末值常量外推尾部 NaN，与"两端缺失保留"
    # 的承诺及 forecast 文案不符，且未被披露）
    s = s.interpolate(method="linear", limit_area="inside")
    tail_nan = int(s.isna().sum())          # 头部/尾部未插值的缺失数
    interpolated = gap_total - tail_nan     # 真实插值数

    meta = {
        "n": int(s.size),
        "dropped_invalid_dates": invalid,
        "merged_duplicates": merged,
        "dup_note": dup_note,
        "interpolated": interpolated,
        "tail_nan": tail_nan,
        "freq": str(freq),
        "utc_note": utc_note,
        "n_before_resample": n_before,
    }
    return s, meta


def _safe_name(name: str) -> str:
    """Windows 文件名清洗（红队裁决 9）：非法字符→_、去首尾空格点、保留名加前缀、限长 40。"""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip(" .") or "all"
    if name.lower() in _WIN_RESERVED:
        name = "_" + name
    return name[:40]


def _cleanup_old_dirs(base_dir: Path, keep_days: int = 30) -> None:
    """删除 base_dir 下超过 keep_days 天的按日期归档目录（外部评审裁决：产物目录必须有清理策略）。

    reports/ 下一切产物子目录（plots/imputed/…）共用；仅识别 %Y%m%d 命名的日目录。
    """
    today = datetime.now().date()
    try:
        for d in Path(base_dir).iterdir():
            if not d.is_dir():
                continue
            try:
                day = datetime.strptime(d.name, "%Y%m%d").date()  # noqa: DTZ007 归档目录名即本地日期，无时区比较
            except ValueError:
                continue
            if (today - day).days > keep_days:
                shutil.rmtree(d, ignore_errors=True)
    except OSError as e:
        # 清理为尽力而为（红队 P2-6）：失败留痕不静默，便于发现产物目录堆积
        logger.warning("30 天归档目录清理失败（不影响本次输出）：%s", e)


def save_plot(fig: Any, name: str) -> str:
    """按附录 D 保存图片，存 reports/plots/YYYYmmdd/ 子目录（红队 C B2：防堆积且不删旧图）。

    文件名 = 清洗后 name_YYYYmmdd_HHMMSS_fff.png（毫秒时间戳防同秒覆盖，红队 B S1）；
    name 由调用方拼装为 "<工具名>_<主列名或all>"；中文字体由模块顶层统一配置。
    返回图片绝对路径字符串（__image__ 字段的值）。
    """
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S_%f")[:-3]
    day_dir = PLOT_DIR / now.strftime("%Y%m%d")
    day_dir.mkdir(parents=True, exist_ok=True)
    _cleanup_old_dirs(PLOT_DIR)
    fname = f"{_safe_name(name)}_{ts}.png"
    out = day_dir / fname
    fig.savefig(out, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return str(out)
